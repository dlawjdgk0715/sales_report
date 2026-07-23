# excel_gen.py
# 별도매출현황 Excel 파일 생성
# 시트 순서: 거래처별요약 → N월 (월별) → Raw_Data

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
import pandas as pd
from config import TOP_GAME_VENDORS, GAME_SUB_ORDER, BIZ_SUB_ORDER, ETC_SUB_ORDER

# ── 색상 팔레트 ─────────────────────────────────────────────────
C_HEADER_GAME = "1F4E79"   # 게임매출 헤더 (진파랑)
C_HEADER_ETC  = "375623"   # 기타수익 헤더 (진초록)
C_HEADER_SVC  = "7030A0"   # 용역수익 헤더 (보라)
C_SUB_GAME    = "D6E4F0"   # 게임매출 소계 배경
C_SUB_ETC     = "E2EFDA"   # 기타수익 소계 배경
C_SUB_SVC     = "EAD7F7"   # 용역수익 소계 배경
C_TOTAL       = "FFF2CC"   # 총합계 배경
C_SECTION_BG  = "F2F2F2"   # 섹션 타이틀 배경
C_ROW_ALT     = "FAFAFA"   # 교차 행 배경

NUM_FMT  = '#,##0'         # 숫자 서식
FONT_NAME = "Arial"

thin = Side(style="thin", color="CCCCCC")
med  = Side(style="medium", color="888888")
THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
MED_BORDER  = Border(left=med,  right=med,  top=med,  bottom=med)


def _font(bold=False, size=10, color="000000"):
    return Font(name=FONT_NAME, bold=bold, size=size, color=color)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _apply_header(ws, row, col, text, bg, fg="FFFFFF", bold=True, size=10):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = _font(bold=bold, size=size, color=fg)
    cell.fill = _fill(bg)
    cell.alignment = _align("center")
    cell.border = THIN_BORDER
    return cell

def _apply_num(ws, row, col, value_or_formula, bg=None):
    cell = ws.cell(row=row, column=col, value=value_or_formula)
    cell.number_format = NUM_FMT
    cell.alignment = _align("right")
    cell.border = THIN_BORDER
    if bg:
        cell.fill = _fill(bg)
    return cell

def _apply_label(ws, row, col, text, bg=None, bold=False, indent=0):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = _font(bold=bold)
    cell.alignment = Alignment(horizontal="left", vertical="center",
                               indent=indent, wrap_text=False)
    cell.border = THIN_BORDER
    if bg:
        cell.fill = _fill(bg)
    return cell


# ── Raw_Data 컬럼 위치 (1-based) ─────────────────────────────────
# 원본 30개 컬럼 뒤에 추가 컬럼
RAW_COL_COUNT = 30          # 원본 컬럼 수
COL_CREDIT = 7              # 대변 (G열)
COL_VENDOR = 11             # 거래처 (K열)
# 추가 컬럼
COL_LABEL   = RAW_COL_COUNT + 1   # 매출구분  (AE)
COL_SUB     = RAW_COL_COUNT + 2   # 세부구분  (AF)
COL_DISPLAY = RAW_COL_COUNT + 3   # 표시명    (AG)
COL_MONTH   = RAW_COL_COUNT + 4   # 월        (AH)
COL_NOTE    = RAW_COL_COUNT + 5   # 비고_매핑 (AI)

COL_DISPLAY_LETTER = get_column_letter(COL_DISPLAY)
COL_MONTH_LETTER   = get_column_letter(COL_MONTH)
COL_CREDIT_LETTER  = get_column_letter(COL_CREDIT)
COL_LABEL_LETTER   = get_column_letter(COL_LABEL)
COL_SUB_LETTER     = get_column_letter(COL_SUB)

RAW_SHEET = "Raw_Data"


def _sumifs_formula(month_str, display_name, category=None, sub=None):
    """
    Raw_Data 참조 SUMIFS 수식 생성
    """
    base = (
        f"SUMIFS({RAW_SHEET}!${COL_CREDIT_LETTER}:${COL_CREDIT_LETTER},"
        f"{RAW_SHEET}!${COL_DISPLAY_LETTER}:${COL_DISPLAY_LETTER},\"{display_name}\","
        f"{RAW_SHEET}!${COL_MONTH_LETTER}:${COL_MONTH_LETTER},\"{month_str}\""
    )
    if category:
        base += f",{RAW_SHEET}!${COL_LABEL_LETTER}:${COL_LABEL_LETTER},\"{category}\""
    if sub:
        base += f",{RAW_SHEET}!${COL_SUB_LETTER}:${COL_SUB_LETTER},\"{sub}\""
    return "=" + base + ")"


# ── 월별 시트 ────────────────────────────────────────────────────

def _write_section_header(ws, row, title, bg_color):
    """섹션 타이틀 행 (■ 게임매출 등)"""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=1, value=f"■ {title}")
    cell.font = _font(bold=True, size=10, color="FFFFFF")
    cell.fill = _fill(bg_color)
    cell.alignment = _align("left")
    cell.border = THIN_BORDER
    for c in range(2, 5):
        ws.cell(row=row, column=c).fill = _fill(bg_color)
        ws.cell(row=row, column=c).border = THIN_BORDER
    return row + 1


def _write_col_headers(ws, row):
    headers = ["구분", "거래처명", "금액", "비고"]
    for i, h in enumerate(headers, 1):
        _apply_header(ws, row, i, h, "404040")
    return row + 1


def build_monthly_sheet(ws, month_str, agg_df, year):
    """
    월별 시트 작성
    agg_df: aggregate_by_month_vendor() 결과
    """
    # 타이틀
    ws.merge_cells("A1:C1")
    t = ws.cell(row=1, column=1, value=f"{year}년 {month_str} 매출현황")
    t.font = _font(bold=True, size=13)
    t.alignment = _align("center")
    unit = ws.cell(row=1, column=4, value="(단위: 원)")
    unit.alignment = _align("right")
    unit.font = _font(size=9, color="666666")

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 6   # 공백

    # 열 너비
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 40

    row = 3
    month_data = agg_df[agg_df["월"] == month_str] if not agg_df.empty else pd.DataFrame()

    # ── 게임매출 ─────────────────────────────────────────────────
    row = _write_section_header(ws, row, "게임매출", C_HEADER_GAME)
    row = _write_col_headers(ws, row)
    hdr_row = row - 1

    sub_start_rows = {}  # sub → first data row
    sub_end_rows   = {}  # sub → last data row

    for sub in GAME_SUB_ORDER:
        sub_vendors = (
            month_data[(month_data["매출구분"] == "게임매출") & (month_data["세부구분"] == sub)]
            if not month_data.empty else pd.DataFrame()
        )
        # 매핑에 있는 모든 벤더 + 실제 데이터에 있는 벤더 합집합
        from config import VENDOR_MAP
        known = {v[0] for v in VENDOR_MAP.get("4100100", {}).values() if v[1] == sub}
        actual_names = set(sub_vendors["표시명"].tolist()) if not sub_vendors.empty else set()
        all_names = sorted(known | actual_names)

        if not all_names:
            continue

        sub_start_rows[sub] = row
        for name in all_names:
            formula = _sumifs_formula(month_str, name, "게임매출", sub)
            note_series = (
                sub_vendors[sub_vendors["표시명"] == name]["비고_매핑"]
                if not sub_vendors.empty else pd.Series(dtype=str)
            )
            note = note_series.iloc[0] if not note_series.empty else ""
            bg = C_ROW_ALT if row % 2 == 0 else None
            _apply_label(ws, row, 1, sub, bg, indent=1)
            _apply_label(ws, row, 2, name, bg)
            _apply_num(ws, row, 3, formula, bg)
            _apply_label(ws, row, 4, note, bg)
            row += 1
        sub_end_rows[sub] = row - 1

    # 소계 행
    subtotal_rows = {}
    for sub in GAME_SUB_ORDER:
        if sub not in sub_start_rows:
            continue
        s, e = sub_start_rows[sub], sub_end_rows[sub]
        c_ref = get_column_letter(3)
        formula = f"=SUM({c_ref}{s}:{c_ref}{e})"
        _apply_label(ws, row, 1, f"{sub} 소계", C_SUB_GAME, bold=True)
        _apply_label(ws, row, 2, "", C_SUB_GAME)
        _apply_num(ws, row, 3, formula, C_SUB_GAME)
        ws.cell(row=row, column=3).font = _font(bold=True)
        _apply_label(ws, row, 4, "", C_SUB_GAME)
        subtotal_rows[sub] = row
        row += 1

    # 게임매출 합계
    game_total_row = row
    if subtotal_rows:
        refs = "+".join(f"C{r}" for r in subtotal_rows.values())
        game_formula = f"={refs}"
    else:
        game_formula = "=0"
    for c in range(1, 5):
        ws.cell(row=row, column=c).fill = _fill(C_SUB_GAME)
        ws.cell(row=row, column=c).border = THIN_BORDER
    _apply_label(ws, row, 1, "게임매출 합계", C_SUB_GAME, bold=True)
    _apply_label(ws, row, 2, "", C_SUB_GAME)
    _apply_num(ws, row, 3, game_formula, C_SUB_GAME)
    ws.cell(row=row, column=3).font = _font(bold=True)
    row += 2  # 공백 1행

    # ── 기타수익 ─────────────────────────────────────────────────
    row = _write_section_header(ws, row, "기타수익", C_HEADER_ETC)
    row = _write_col_headers(ws, row)

    etc_sub_rows = {}
    for sub in ETC_SUB_ORDER:
        sub_vendors = (
            month_data[(month_data["매출구분"] == "기타수익") & (month_data["세부구분"] == sub)]
            if not month_data.empty else pd.DataFrame()
        )
        from config import VENDOR_MAP as VM
        known = {v[0] for v in VM.get("4100500", {}).values() if v[1] == sub}
        actual_names = set(sub_vendors["표시명"].tolist()) if not sub_vendors.empty else set()
        all_names = sorted(known | actual_names)

        if not all_names:
            continue

        s_row = row
        for name in all_names:
            formula = _sumifs_formula(month_str, name, "기타수익", sub)
            note_series = (
                sub_vendors[sub_vendors["표시명"] == name]["비고_매핑"]
                if not sub_vendors.empty else pd.Series(dtype=str)
            )
            note = note_series.iloc[0] if not note_series.empty else ""
            bg = C_ROW_ALT if row % 2 == 0 else None
            _apply_label(ws, row, 1, sub, bg, indent=1)
            _apply_label(ws, row, 2, name, bg)
            _apply_num(ws, row, 3, formula, bg)
            _apply_label(ws, row, 4, note, bg)
            row += 1
        etc_sub_rows[sub] = (s_row, row - 1)

    # 기타수익 소계
    etc_total_row = row
    if etc_sub_rows:
        etc_formula = "=SUM(" + ",".join(f"C{s}:C{e}" for s, e in etc_sub_rows.values()) + ")"
    else:
        etc_formula = "=0"
    for c in range(1, 5):
        ws.cell(row=row, column=c).fill = _fill(C_SUB_ETC)
        ws.cell(row=row, column=c).border = THIN_BORDER
    _apply_label(ws, row, 1, "기타수익 합계", C_SUB_ETC, bold=True)
    _apply_label(ws, row, 2, "", C_SUB_ETC)
    _apply_num(ws, row, 3, etc_formula, C_SUB_ETC)
    ws.cell(row=row, column=3).font = _font(bold=True)
    row += 2

    # ── 용역수익 ─────────────────────────────────────────────────
    row = _write_section_header(ws, row, "용역수익", C_HEADER_SVC)
    row = _write_col_headers(ws, row)

    svc_vendors = (
        month_data[month_data["매출구분"] == "용역수익"]
        if not month_data.empty else pd.DataFrame()
    )
    from config import VENDOR_MAP as VM2
    known_svc = {v[0] for v in VM2.get("4100300", {}).values()}
    actual_svc = set(svc_vendors["표시명"].tolist()) if not svc_vendors.empty else set()
    all_svc = sorted(known_svc | actual_svc)

    svc_start = row
    for name in all_svc:
        formula = _sumifs_formula(month_str, name, "용역수익")
        note_series = (
            svc_vendors[svc_vendors["표시명"] == name]["비고_매핑"]
            if not svc_vendors.empty else pd.Series(dtype=str)
        )
        note = note_series.iloc[0] if not note_series.empty else ""
        bg = C_ROW_ALT if row % 2 == 0 else None
        _apply_label(ws, row, 1, "경영관리수익", bg, indent=1)
        _apply_label(ws, row, 2, name, bg)
        _apply_num(ws, row, 3, formula, bg)
        _apply_label(ws, row, 4, note, bg)
        row += 1
    svc_end = row - 1

    # 용역수익 합계
    svc_total_row = row
    svc_formula = f"=SUM(C{svc_start}:C{svc_end})" if all_svc else "=0"
    for c in range(1, 5):
        ws.cell(row=row, column=c).fill = _fill(C_SUB_SVC)
        ws.cell(row=row, column=c).border = THIN_BORDER
    _apply_label(ws, row, 1, "용역수익 합계", C_SUB_SVC, bold=True)
    _apply_label(ws, row, 2, "", C_SUB_SVC)
    _apply_num(ws, row, 3, svc_formula, C_SUB_SVC)
    ws.cell(row=row, column=3).font = _font(bold=True)
    row += 2

    # ── 총합계 ───────────────────────────────────────────────────
    for c in range(1, 5):
        ws.cell(row=row, column=c).fill = _fill(C_TOTAL)
        ws.cell(row=row, column=c).border = MED_BORDER
    _apply_label(ws, row, 1, "총  합  계", C_TOTAL, bold=True)
    ws.cell(row=row, column=1).font = _font(bold=True, size=11)
    ws.cell(row=row, column=1).alignment = _align("center")
    _apply_label(ws, row, 2, "", C_TOTAL)
    total_formula = f"=C{game_total_row}+C{etc_total_row}+C{svc_total_row}"
    _apply_num(ws, row, 3, total_formula, C_TOTAL)
    ws.cell(row=row, column=3).font = _font(bold=True, size=11)

    # 창 고정 (헤더 고정)
    ws.freeze_panes = "A3"


# ── 거래처별 요약 시트 ─────────────────────────────────────────────

def build_summary_sheet(ws, months, agg_df, year):
    """
    거래처별 요약 시트: 게임매출 상위 거래처 × 월별
    months: ["1월", "2월", ...] 오름차순
    """
    # 타이틀
    col_count = len(months) + 2
    t = ws.cell(row=1, column=1, value=f"{year}년 컴투스 매출현황 (게임매출 주요 거래처)")
    t.font = _font(bold=True, size=13)
    t.alignment = _align("center")

    ws.cell(row=2, column=col_count, value="(단위: 백만원)").alignment = _align("right")

    # 헤더 행
    row = 3
    _apply_header(ws, row, 1, "분류", C_HEADER_GAME)
    _apply_header(ws, row, 2, "거래처", C_HEADER_GAME)
    for i, m in enumerate(months, 3):
        _apply_header(ws, row, i, m, C_HEADER_GAME)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 20
    for i in range(3, col_count + 1):
        ws.column_dimensions[get_column_letter(i)].width = 13

    row = 4
    # 게임매출 상위 7개 + 기타 + 합계
    vendor_rows = {}

    for vendor in TOP_GAME_VENDORS:
        bg = C_ROW_ALT if row % 2 == 0 else None
        _apply_label(ws, row, 1, "게임매출", bg)
        _apply_label(ws, row, 2, vendor, bg)
        for i, m in enumerate(months, 3):
            # 백만원 단위 SUMIFS ÷ 1,000,000
            f = f"=IFERROR({_sumifs_formula(m, vendor, '게임매출')[1:]}/1000000,0)"
            _apply_num(ws, row, i, f, bg)
        vendor_rows[vendor] = row
        row += 1

    # 기타 (합계 - 상위7)
    _apply_label(ws, row, 1, "게임매출", C_ROW_ALT)
    _apply_label(ws, row, 2, "기타", C_ROW_ALT)
    other_row = row
    for i, m in enumerate(months, 3):
        # 게임매출 전체 - 상위7 합
        total_f = f"SUMIFS({RAW_SHEET}!${COL_CREDIT_LETTER}:${COL_CREDIT_LETTER},{RAW_SHEET}!${COL_MONTH_LETTER}:${COL_MONTH_LETTER},\"{m}\",{RAW_SHEET}!${COL_LABEL_LETTER}:${COL_LABEL_LETTER},\"게임매출\")"
        known_refs = "+".join(
            f"SUMIFS({RAW_SHEET}!${COL_CREDIT_LETTER}:${COL_CREDIT_LETTER},"
            f"{RAW_SHEET}!${COL_DISPLAY_LETTER}:${COL_DISPLAY_LETTER},\"{v}\","
            f"{RAW_SHEET}!${COL_MONTH_LETTER}:${COL_MONTH_LETTER},\"{m}\","
            f"{RAW_SHEET}!${COL_LABEL_LETTER}:${COL_LABEL_LETTER},\"게임매출\")"
            for v in TOP_GAME_VENDORS
        )
        f = f"=IFERROR(({total_f}-({known_refs}))/1000000,0)"
        _apply_num(ws, row, i, f, C_ROW_ALT)
    row += 1

    # 합계 행
    for c in range(1, col_count + 1):
        ws.cell(row=row, column=c).fill = _fill(C_SUB_GAME)
        ws.cell(row=row, column=c).border = THIN_BORDER
    ws.cell(row=row, column=1, value="게임매출").font = _font(bold=True)
    ws.cell(row=row, column=1).fill = _fill(C_SUB_GAME)
    ws.cell(row=row, column=1).alignment = _align("left")
    ws.cell(row=row, column=2, value="합계").font = _font(bold=True)
    ws.cell(row=row, column=2).fill = _fill(C_SUB_GAME)
    ws.cell(row=row, column=2).alignment = _align("left")
    game_total_summary_row = row
    for i, m in enumerate(months, 3):
        f = (f"=IFERROR(SUMIFS({RAW_SHEET}!${COL_CREDIT_LETTER}:${COL_CREDIT_LETTER},"
             f"{RAW_SHEET}!${COL_MONTH_LETTER}:${COL_MONTH_LETTER},\"{m}\","
             f"{RAW_SHEET}!${COL_LABEL_LETTER}:${COL_LABEL_LETTER},\"게임매출\")/1000000,0)")
        _apply_num(ws, row, i, f, C_SUB_GAME)
        ws.cell(row=row, column=i).font = _font(bold=True)
    row += 1

    # 기타수익 행
    for c in range(1, col_count + 1):
        ws.cell(row=row, column=c).border = THIN_BORDER
    _apply_label(ws, row, 1, "기타수익")
    _apply_label(ws, row, 2, "합계")
    for i, m in enumerate(months, 3):
        f = (f"=IFERROR(SUMIFS({RAW_SHEET}!${COL_CREDIT_LETTER}:${COL_CREDIT_LETTER},"
             f"{RAW_SHEET}!${COL_MONTH_LETTER}:${COL_MONTH_LETTER},\"{m}\","
             f"{RAW_SHEET}!${COL_LABEL_LETTER}:${COL_LABEL_LETTER},\"기타수익\")/1000000,0)")
        _apply_num(ws, row, i, f)
    etc_summary_row = row
    row += 1

    # 용역수익 행
    for c in range(1, col_count + 1):
        ws.cell(row=row, column=c).border = THIN_BORDER
    _apply_label(ws, row, 1, "용역수익")
    _apply_label(ws, row, 2, "합계")
    for i, m in enumerate(months, 3):
        f = (f"=IFERROR(SUMIFS({RAW_SHEET}!${COL_CREDIT_LETTER}:${COL_CREDIT_LETTER},"
             f"{RAW_SHEET}!${COL_MONTH_LETTER}:${COL_MONTH_LETTER},\"{m}\","
             f"{RAW_SHEET}!${COL_LABEL_LETTER}:${COL_LABEL_LETTER},\"용역수익\")/1000000,0)")
        _apply_num(ws, row, i, f)
    svc_summary_row = row
    row += 1

    # 총합계 행
    for c in range(1, col_count + 1):
        ws.cell(row=row, column=c).fill = _fill(C_TOTAL)
        ws.cell(row=row, column=c).border = MED_BORDER
    ws.cell(row=row, column=1, value="합계").font = _font(bold=True, size=11)
    ws.cell(row=row, column=1).fill = _fill(C_TOTAL)
    ws.cell(row=row, column=1).alignment = _align("center")
    ws.cell(row=row, column=2, value="총계").font = _font(bold=True)
    ws.cell(row=row, column=2).fill = _fill(C_TOTAL)
    ws.cell(row=row, column=2).alignment = _align("left")
    for i in range(3, col_count + 1):
        f = f"={get_column_letter(i)}{game_total_summary_row}+{get_column_letter(i)}{etc_summary_row}+{get_column_letter(i)}{svc_summary_row}"
        _apply_num(ws, row, i, f, C_TOTAL)
        ws.cell(row=row, column=i).font = _font(bold=True, size=11)

    ws.freeze_panes = "C4"


# ── Raw_Data 시트 ─────────────────────────────────────────────────

def build_raw_data_sheet(ws, raw_df):
    """
    Raw_Data 시트: 원본 데이터 + 매핑 컬럼
    """
    if raw_df is None or raw_df.empty:
        ws.cell(row=1, column=1, value="데이터 없음")
        return

    # 컬럼 순서: 원본 30개 + 추가 5개 (SUMIFS 수식이 고정 열 위치 참조하므로 절대 변경 불가)
    extra_cols = ["매출구분", "세부구분", "표시명", "월", "비고_매핑"]
    original_cols = [c for c in raw_df.columns if c not in extra_cols]
    all_cols = original_cols + [c for c in extra_cols if c in raw_df.columns]
    raw_df = raw_df[all_cols]

    # 헤더
    for c_idx, col in enumerate(all_cols, 1):
        cell = ws.cell(row=1, column=c_idx, value=col)
        is_extra = col in extra_cols
        cell.font = _font(bold=True, color="FFFFFF")
        cell.fill = _fill("1F4E79" if is_extra else "404040")
        cell.alignment = _align("center")
        cell.border = THIN_BORDER

    # 데이터 행
    for r_idx, row_data in enumerate(raw_df.itertuples(index=False), 2):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = THIN_BORDER
            cell.font = _font()
            if isinstance(val, (int, float)) and c_idx == COL_CREDIT:
                cell.number_format = NUM_FMT
                cell.alignment = _align("right")

    # 열 너비
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["D"].width = 14   # 회계일
    ws.column_dimensions["I"].width = 40   # 적요
    ws.column_dimensions["K"].width = 36   # 거래처
    for c_idx in range(RAW_COL_COUNT + 1, len(all_cols) + 1):
        ws.column_dimensions[get_column_letter(c_idx)].width = 18

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(all_cols))}1"


# ── 조회 시트 ────────────────────────────────────────────────────

def build_lookup_sheet(ws, wb, months, agg_df, raw_row_count=5000):
    """
    FILTER 함수 기반 Raw_Data 검색 시트.
    B2=월, B3=거래처(표시명), B4=매출구분 자유 입력 → 행7부터 결과 자동 출력.
    - 쉼표(,)로 다중값 입력 가능: 예) "5월,6월" → 5월 또는 6월 행 모두 표시
    - (전체) 입력 또는 빈칸이면 해당 조건 미적용
    - 드롭다운 없음 — 자유 텍스트 입력 방식
    - 숨김 시트(__Lists__)는 유지 (향후 드롭다운 추가 시 활용 가능)
    """

    # ── 숨김 목록 시트 생성 (참고용, 드롭다운은 미적용) ──────────
    LIST_SHEET = "__Lists__"
    wl = wb.create_sheet(LIST_SHEET)
    wl.sheet_state = "hidden"

    # A열: 월 목록
    wl.cell(row=1, column=1, value="(전체)")
    for i, m in enumerate(months, 2):
        wl.cell(row=i, column=1, value=m)

    # B열: 표시명 목록
    if not agg_df.empty and "표시명" in agg_df.columns:
        vendor_list = ["(전체)"] + sorted(agg_df["표시명"].dropna().unique().tolist())
    else:
        vendor_list = ["(전체)"]
    for i, v in enumerate(vendor_list, 1):
        wl.cell(row=i, column=2, value=v)

    # C열: 매출구분 목록
    cats = ["(전체)", "게임매출", "기타수익", "용역수익"]
    for i, c in enumerate(cats, 1):
        wl.cell(row=i, column=3, value=c)

    # ── 열 너비 ─────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 28
    ws.column_dimensions["I"].width = 44
    ws.column_dimensions["J"].width = 14

    # ── 타이틀 ───────────────────────────────────────────────────
    ws.merge_cells("A1:J1")
    t = ws.cell(row=1, column=1, value="Raw_Data 상세 조회")
    t.font = Font(name=FONT_NAME, bold=True, size=13, color="FFFFFF")
    t.fill = _fill("1F4E79")
    t.alignment = _align("left")
    ws.row_dimensions[1].height = 22

    # ── 필터 입력 셀 ─────────────────────────────────────────────
    FILTER_BG   = "EBF3FF"
    LABEL_COLOR = "1F4E79"
    INPUT_BORDER = Border(
        left=Side(style="medium", color="1A73E8"),
        right=Side(style="medium", color="1A73E8"),
        top=Side(style="medium", color="1A73E8"),
        bottom=Side(style="medium", color="1A73E8"),
    )

    filter_defs = [
        (2, "월",           "(전체)"),
        (3, "거래처(표시명)", "(전체)"),
        (4, "매출구분",      "(전체)"),
    ]
    for row, label, default in filter_defs:
        lc = ws.cell(row=row, column=1, value=label)
        lc.font      = _font(bold=True, color=LABEL_COLOR)
        lc.fill      = _fill(FILTER_BG)
        lc.alignment = _align("right")
        lc.border    = THIN_BORDER
        vc = ws.cell(row=row, column=2, value=default)
        vc.font      = _font(bold=True)
        vc.fill      = _fill("FFFFFF")
        vc.border    = INPUT_BORDER
        vc.alignment = _align("left")
        ws.row_dimensions[row].height = 18

    # 힌트 (C2:J4 병합)
    ws.merge_cells("C2:J4")
    hint = ws.cell(row=2, column=3,
                   value="드롭다운 또는 쉼표(,)로 다중입력 (예: 5월,6월)  |  (전체) 선택시 해당 조건 미적용")
    hint.font = _font(size=9, color="666666")

    # ── 구분선 ───────────────────────────────────────────────────
    ws.row_dimensions[5].height = 4
    for c in range(1, 11):
        ws.cell(row=5, column=c).fill = _fill("1A73E8")

    # ── 결과 컬럼 헤더 (row 6) ───────────────────────────────────
    result_cols = ["No", "거래처", "회계일", "월", "대변", "매출구분", "세부구분", "표시명", "적요", "비고_매핑"]
    for ci, h in enumerate(result_cols, 1):
        _apply_header(ws, 6, ci, h, "2C5F8A")
    ws.row_dimensions[6].height = 18

    # ── FILTER 수식 (A7) ─────────────────────────────────────────
    # 다중값(쉼표 구분) 지원:
    #   ISNUMBER(SEARCH(","&값&",", ","&SUBSTITUTE(B2," ","")&","))
    #   → "5월,6월" 입력 시 5월·6월 모두 매칭
    # (전체) 또는 빈칸이면 해당 조건 전체 TRUE (미적용)
    # 매출구분(B4)은 단일값 완전 일치만 지원 (3개뿐이라 다중 불필요)
    # 실제 Raw_Data 행수 + 여유분(500) — 향후 누적 업데이트 시 범위 자동 확장
    MAX_ROW = max(raw_row_count + 500, 1000)
    R  = "Raw_Data"
    AH = COL_MONTH_LETTER
    AG = COL_DISPLAY_LETTER
    AE = COL_LABEL_LETTER
    AF = COL_SUB_LETTER
    AI = get_column_letter(COL_NOTE)

    # 월 조건: (전체) 또는 SEARCH 다중 매칭
    month_cond = (
        f'(($B$2="(전체)")'
        f'+ISNUMBER(SEARCH(","&{R}!${AH}$2:${AH}${MAX_ROW}&",",","&SUBSTITUTE($B$2," ","")&",")))'
    )
    # 거래처 조건: (전체) 또는 SEARCH 다중 매칭
    vendor_cond = (
        f'(($B$3="(전체)")'
        f'+ISNUMBER(SEARCH(","&{R}!${AG}$2:${AG}${MAX_ROW}&",",","&SUBSTITUTE($B$3," ","")&",")))'
    )
    # 매출구분 조건: (전체) 또는 완전 일치
    cat_cond = (
        f'(($B$4="(전체)")+({R}!${AE}$2:${AE}${MAX_ROW}=$B$4))'
    )

    cond = f'{month_cond}*{vendor_cond}*{cat_cond}'

    cols = (
        f'{R}!$A$2:$A${MAX_ROW},'
        f'{R}!$K$2:$K${MAX_ROW},'
        f'{R}!$D$2:$D${MAX_ROW},'
        f'{R}!${AH}$2:${AH}${MAX_ROW},'
        f'{R}!$G$2:$G${MAX_ROW},'
        f'{R}!${AE}$2:${AE}${MAX_ROW},'
        f'{R}!${AF}$2:${AF}${MAX_ROW},'
        f'{R}!${AG}$2:${AG}${MAX_ROW},'
        f'{R}!$I$2:$I${MAX_ROW},'
        f'{R}!${AI}$2:${AI}${MAX_ROW}'
    )
    # _xlfn._xlws. 접두어: 일부 Excel 버전에서 동적 배열 FILTER 인식에 필요
    formula = (
        f'_xlfn._xlws.FILTER('
        f'CHOOSE({{1,2,3,4,5,6,7,8,9,10}},{cols}),'
        f'{cond},""'
        f')'
    )

    # 자리 표시 더미값 (XML 주입 위치 식별용)
    ws.cell(row=7, column=1, value="__FILTER_FORMULA__")
    ws.cell(row=7, column=1).font = _font(color="1A1A7E")

    ws.freeze_panes = "A7"

    # __Lists__ 참조 범위 (x14 드롭다운 주입용)
    LIST_SHEET = "__Lists__"
    month_ref  = f"'{LIST_SHEET}'!$A$1:$A${1+len(months)}"
    if not agg_df.empty and "표시명" in agg_df.columns:
        n_vendors = 1 + len(agg_df["표시명"].dropna().unique())
    else:
        n_vendors = 1
    vendor_ref = f"'{LIST_SHEET}'!$B$1:$B${n_vendors}"
    cat_ref    = f"'{LIST_SHEET}'!$C$1:$C$4"

    return formula, month_ref, vendor_ref, cat_ref


# ── FILTER 수식 XML 직접 주입 ─────────────────────────────────────

def _inject_filter_formula(xlsx_path, sheet_name, formula):
    """
    openpyxl이 저장한 xlsx를 zip으로 열어 해당 시트 XML에서
    더미 문자열 셀을 FILTER 수식 셀로 교체.
    openpyxl의 한글/특수문자 인코딩 문제를 완전히 우회.
    """
    import zipfile, shutil, os, re, tempfile

    tmp = xlsx_path + ".tmp"
    with zipfile.ZipFile(xlsx_path, "r") as zin:
        # workbook.xml에서 sheet_name의 파일명 확인
        wb_xml = zin.read("xl/workbook.xml").decode("utf-8")
        rels_xml = zin.read("xl/_rels/workbook.xml.rels").decode("utf-8")

        # rId → Target 매핑 (속성 순서 무관하게 처리)
        rid_to_target = {}
        for m in re.finditer(r'<Relationship\b([^>]+)>', rels_xml):
            attrs = m.group(1)
            id_m  = re.search(r'\bId="([^"]+)"', attrs)
            tgt_m = re.search(r'\bTarget="([^"]+)"', attrs)
            if not id_m or not tgt_m:
                continue
            rid, tgt = id_m.group(1), tgt_m.group(1)
            # /xl/worksheets/sheetN.xml  또는  worksheets/sheetN.xml 모두 → xl/worksheets/sheetN.xml
            tgt = tgt.lstrip("/")
            if not tgt.startswith("xl/"):
                tgt = "xl/" + tgt
            rid_to_target[rid] = tgt

        # 시트명 → 파일 경로
        name_to_file = {}
        for m in re.finditer(r'<sheet\b[^>]+\bname="([^"]+)"[^>]+\br:id="([^"]+)"', wb_xml):
            sname, rid = m.group(1), m.group(2)
            if rid in rid_to_target:
                name_to_file[sname] = rid_to_target[rid]

        target_file = name_to_file.get(sheet_name)
        if not target_file:
            return  # 시트 못 찾으면 건너뜀

        with zipfile.ZipFile(tmp, "w", compression=zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == target_file:
                    xml = data.decode("utf-8")
                    # 더미 문자열 "__FILTER_FORMULA__" 가 있는 <c> 태그를 수식 셀로 교체
                    # 패턴: <c r="A7" ...><v>...__FILTER_FORMULA__...</v></c>
                    # → <c r="A7" t="str"><f>IFERROR(FILTER(...))</f></c>
                    # inlineStr / shared / 일반 등 모든 타입의 A7 셀 교체
                    xml = re.sub(
                        r'<c r="A7"[^>]*(?:/>|>.*?</c>)',
                        f'<c r="A7"><f>{formula}</f></c>',
                        xml,
                        flags=re.DOTALL
                    )
                    data = xml.encode("utf-8")
                zout.writestr(item, data)

    os.replace(tmp, xlsx_path)


def _inject_x14_datavalidation(xlsx_path, sheet_name, month_ref, vendor_ref, cat_ref):
    """
    openpyxl은 x14:dataValidations(확장 네임스페이스) 를 쓰지 않으므로
    xlsx zip을 직접 수정해 B2/B3/B4 드롭다운을 __Lists__ 범위로 연결.
    이 방식은 드롭다운 선택 + 자유 입력(쉼표 다중값) 동시 지원.
    """
    import zipfile, os, re

    tmp = xlsx_path + ".tmp"
    x14_block = (
        '<extLst>'
        '<ext uri="{CCE6A557-97BC-4b89-ADB6-D9C93CAAB3DF}" '
        'xmlns:x14="http://schemas.microsoft.com/office/spreadsheetml/2009/9/main">'
        '<x14:dataValidations count="3" '
        'xmlns:xm="http://schemas.microsoft.com/office/excel/2006/main">'
        f'<x14:dataValidation type="list">'
        f'<x14:formula1><xm:f>{month_ref}</xm:f></x14:formula1>'
        f'<xm:sqref>B2</xm:sqref>'
        f'</x14:dataValidation>'
        f'<x14:dataValidation type="list">'
        f'<x14:formula1><xm:f>{vendor_ref}</xm:f></x14:formula1>'
        f'<xm:sqref>B3</xm:sqref>'
        f'</x14:dataValidation>'
        f'<x14:dataValidation type="list">'
        f'<x14:formula1><xm:f>{cat_ref}</xm:f></x14:formula1>'
        f'<xm:sqref>B4</xm:sqref>'
        f'</x14:dataValidation>'
        '</x14:dataValidations>'
        '</ext>'
        '</extLst>'
    )

    # 시트 파일 경로 찾기
    with zipfile.ZipFile(xlsx_path, "r") as zin:
        wb_xml   = zin.read("xl/workbook.xml").decode("utf-8")
        rels_xml = zin.read("xl/_rels/workbook.xml.rels").decode("utf-8")

        rid_to_target = {}
        for m in re.finditer(r'<Relationship\b([^>]+)>', rels_xml):
            attrs = m.group(1)
            id_m  = re.search(r'\bId="([^"]+)"', attrs)
            tgt_m = re.search(r'\bTarget="([^"]+)"', attrs)
            if id_m and tgt_m:
                tgt = tgt_m.group(1).lstrip("/")
                if not tgt.startswith("xl/"): tgt = "xl/" + tgt
                rid_to_target[id_m.group(1)] = tgt

        name_to_file = {}
        for m in re.finditer(r'<sheet\b[^>]+name="([^"]+)"[^>]+r:id="([^"]+)"', wb_xml):
            sname, rid = m.group(1), m.group(2)
            if rid in rid_to_target:
                name_to_file[sname] = rid_to_target[rid]

        target_file = name_to_file.get(sheet_name)
        if not target_file:
            return

        with zipfile.ZipFile(tmp, "w", compression=zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == target_file:
                    xml = data.decode("utf-8")
                    # 기존 extLst 제거 후 </worksheet> 직전에 삽입
                    xml = re.sub(r'<extLst>.*?</extLst>', '', xml, flags=re.DOTALL)
                    xml = xml.replace('</worksheet>', x14_block + '</worksheet>')
                    data = xml.encode("utf-8")
                zout.writestr(item, data)

    os.replace(tmp, xlsx_path)


def _fix_formula_encoding(xlsx_path):
    """
    openpyxl이 저장한 xlsx의 모든 시트 XML에서
    수식 태그(<f>...</f>) 안의 XML 숫자 엔티티(&#XXXXX;)를
    실제 유니코드 문자로 교체.

    예: "1&#50900;" → "1월", "&#44172;&#51076;&#47588;&#52636;" → "게임매출"

    SUMIFS/VLOOKUP 등의 한글 문자열 기준값이 제대로 매칭되도록 수정.
    """
    import zipfile, os, re, html

    tmp = xlsx_path + ".tmp"

    def decode_entities_in_formulas(xml_text):
        # <f>...</f> 블록 내의 &#숫자; 엔티티만 디코딩
        def replacer(m):
            return "<f>" + html.unescape(m.group(1)) + "</f>"
        return re.sub(r'<f>(.*?)</f>', replacer, xml_text, flags=re.DOTALL)

    with zipfile.ZipFile(xlsx_path, "r") as zin:
        with zipfile.ZipFile(tmp, "w", compression=zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                # 시트 XML 파일만 처리
                if item.filename.startswith("xl/worksheets/sheet") and item.filename.endswith(".xml"):
                    xml = data.decode("utf-8")
                    fixed = decode_entities_in_formulas(xml)
                    if fixed != xml:
                        data = fixed.encode("utf-8")
                zout.writestr(item, data)

    os.replace(tmp, xlsx_path)


# ── 메인 생성 함수 ────────────────────────────────────────────────

def generate_excel(raw_df, agg_df, output_path):
    """
    최종 Excel 파일 생성.
    raw_df: Raw_Data (매핑 컬럼 포함)
    agg_df: aggregate_by_month_vendor() 결과
    output_path: 저장 경로
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # 기본 Sheet 제거

    # 연도 추출
    months_in_data = sorted(
        agg_df["월"].dropna().unique().tolist(),
        key=lambda x: int(x.replace("월", ""))
    ) if not agg_df.empty else []

    year = 2026
    if not raw_df.empty and "회계일" in raw_df.columns:
        dates = pd.to_datetime(raw_df["회계일"], errors="coerce").dropna()
        if not dates.empty:
            year = dates.dt.year.mode()[0]

    # 1. 거래처별 요약 (첫 번째 시트)
    ws_sum = wb.create_sheet("거래처별요약")
    build_summary_sheet(ws_sum, months_in_data, agg_df, year)

    # 2. 월별 시트
    for m in months_in_data:
        ws_m = wb.create_sheet(m)
        build_monthly_sheet(ws_m, m, agg_df, year)

    # 3. 검색 시트 (레이아웃만, 수식은 XML 후처리로 주입)
    ws_lkp = wb.create_sheet("검색")
    lookup_formula, month_ref, vendor_ref, cat_ref = build_lookup_sheet(
        ws_lkp, wb, months_in_data, agg_df,
        raw_row_count=len(raw_df)   # 실제 행수 전달 → FILTER 범위 동적 설정
    )

    # 4. Raw_Data (마지막)
    ws_raw = wb.create_sheet(RAW_SHEET)
    build_raw_data_sheet(ws_raw, raw_df)

    wb.save(output_path)

    # 5. FILTER 수식을 XML에 직접 주입 (openpyxl 인코딩 우회)
    _inject_filter_formula(output_path, "검색", lookup_formula)

    # 5b. x14:dataValidations (드롭다운) XML 주입 — openpyxl 미지원 확장 네임스페이스
    _inject_x14_datavalidation(output_path, "검색", month_ref, vendor_ref, cat_ref)

    # 6. 모든 시트의 수식에서 한글 XML 엔티티 디코딩
    #    (SUMIFS 등 월별/요약 시트 수식의 한글 기준값 매칭 오류 수정)
    _fix_formula_encoding(output_path)

    return output_path
