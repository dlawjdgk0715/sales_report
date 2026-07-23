# processor.py
# grid 파일 읽기 → 필터링 → 매핑 적용 → 통합 DataFrame 반환

import pandas as pd
from config import VENDOR_MAP, ACCOUNT_LABEL, NOTE_SUB_RULES

EXCLUDE_NAMES = {"전일이월", "월계", "누계"}


def load_grid_files(file_paths):
    """
    여러 grid 파일을 읽어 하나의 DataFrame으로 합칩니다.
    각 파일에서 불필요한 행(전일이월, 월계, 누계) 및 거래처 없는 행 제거.
    """
    frames = []
    for path in file_paths:
        df = pd.read_excel(path, dtype={"계정코드": str, "거래처코드": str})
        # 헤더 정리
        df.columns = [str(c).strip() for c in df.columns]
        # 실제 거래 행만 유지
        df = df[
            ~df["계정명"].isin(EXCLUDE_NAMES) &
            df["거래처"].notna() &
            df["거래처"].str.strip().ne("")
        ].copy()
        frames.append(df)

    if not frames:
        return pd.DataFrame()

    combined = pd.concat(frames, ignore_index=True)
    return combined


def apply_mapping(df, extra_map=None):
    """
    DataFrame에 매핑 정보(매출구분, 세부구분, 표시명, 월)를 추가합니다.
    extra_map: {계정코드: {거래처명: (표시명, 세부구분, 비고)}} — 웹앱에서 추가된 신규 분류
    """
    merged_map = {}
    for acct, mapping in VENDOR_MAP.items():
        merged_map[acct] = dict(mapping)
    if extra_map:
        for acct, mapping in extra_map.items():
            if acct not in merged_map:
                merged_map[acct] = {}
            merged_map[acct].update(mapping)

    # 계정코드 기준 매출구분
    df["매출구분"] = df["계정코드"].map(ACCOUNT_LABEL).fillna("기타")

    # 거래처 + 계정코드 기준 세부구분 / 표시명 / 비고
    def _map_row(row):
        acct   = str(row["계정코드"])
        vendor = str(row["거래처"]).strip()
        note_text = str(row.get("적요", "") or "").strip()

        # ── 1. 적요 기반 분류 룰 우선 체크 ─────────────────────────
        note_rule = NOTE_SUB_RULES.get(acct, {}).get(vendor)
        if note_rule:
            for keywords, display, sub, note in note_rule["rules"]:
                if any(kw in note_text for kw in keywords):
                    return pd.Series({"표시명": display, "세부구분": sub, "비고_매핑": note})
            # 키워드 미매칭 → default
            default = note_rule["default"]
            if default is None:
                # 사용자 확인 필요: 표시명은 vendor, 세부구분은 "확인필요"로 표시
                return pd.Series({"표시명": vendor, "세부구분": "확인필요", "비고_매핑": f"[적요] {note_text[:80]}"})
            return pd.Series({"표시명": default[0], "세부구분": default[1], "비고_매핑": default[2]})

        # ── 2. 일반 VENDOR_MAP ────────────────────────────────────
        acct_map = merged_map.get(acct, {})
        info = acct_map.get(vendor)
        if info:
            return pd.Series({"표시명": info[0], "세부구분": info[1], "비고_매핑": info[2]})
        return pd.Series({"표시명": vendor, "세부구분": "미분류", "비고_매핑": ""})

    mapped = df.apply(_map_row, axis=1)
    df = pd.concat([df, mapped], axis=1)

    # 월 컬럼 (예: "6월")
    df["회계일"] = pd.to_datetime(df["회계일"], errors="coerce")
    df["월"] = df["회계일"].dt.month.astype("Int64").astype(str) + "월"

    # 대변 숫자형 보장
    df["대변"] = pd.to_numeric(df["대변"], errors="coerce").fillna(0)

    return df


def find_unmapped_vendors(df):
    """
    매핑 미완료 거래처 목록 반환.
    - "미분류": 매핑 테이블에 없는 신규 거래처
    - "확인필요": 적요 기반 분류 룰이 있으나 키워드 미매칭 → 사용자 확인 필요
    Returns list of (계정코드, 거래처명, 매출구분, 세부구분, 비고_매핑) tuples.
    """
    mask = df["세부구분"].isin(["미분류", "확인필요"])
    cols = ["계정코드", "거래처", "매출구분", "세부구분", "비고_매핑"]
    unmapped = df[mask][cols].drop_duplicates(subset=["계정코드", "거래처"])
    return unmapped.values.tolist()


def aggregate_by_month_vendor(df):
    """
    월 × 표시명 × 매출구분 × 세부구분 기준으로 대변 합산.
    Returns: pivot-ready DataFrame
    """
    agg = (
        df.groupby(["월", "매출구분", "세부구분", "표시명", "비고_매핑"], dropna=False)["대변"]
        .sum()
        .reset_index()
    )
    return agg



# 원본 grid 파일의 30개 컬럼 순서 (SUMIFS 수식이 이 순서를 가정함)
GRID_COLUMNS = [
    "No", "계정코드", "계정명", "회계일", "승인번호", "차변", "대변", "잔액",
    "적요", "거래처코드", "거래처", "비용센터코드", "비용센터", "거래환종",
    "거래금액", "환율", "프로젝트", "프로젝트명", "증빙", "예산단위", "예산계정",
    "회계단위", "작성부서", "작성자", "전표번호", "순번", "메뉴",
    "전표유형코드", "전표유형명", "계정유형",
]
# COL_CREDIT=7(G), 전표번호=25, 순번=26 이 위 순서에 맞아야 SUMIFS가 정상 작동

def _make_legacy_row(vendor, acct, amount, cat, sub_cat, display, month_str, note, seq):
    """
    원본 30컬럼 + 5 추가컬럼 구조의 딕셔너리 반환.
    대변=7번째(G), 거래처=11번째(K), 전표번호·순번도 올바른 위치에 배치.
    """
    row = {c: None for c in GRID_COLUMNS}
    row["계정코드"] = acct
    row["거래처"]   = vendor
    row["대변"]     = amount
    # 결정론적 중복키: 동일 파일 재업로드 시 dedup 처리됨
    row["전표번호"] = f"LEGACY_{month_str}_{vendor}_{cat}"
    row["순번"]     = seq
    # 추가 컬럼 (apply_mapping이 이미 붙여준 것처럼 직접 세팅)
    row["매출구분"] = cat
    row["세부구분"] = sub_cat
    row["표시명"]   = display
    row["월"]       = month_str
    row["비고_매핑"] = note
    return row


def read_legacy_file(file_path):
    """
    구버전 별도매출현황 파일(N월 / N월-1 시트 구조)을 읽어
    현재 형식과 호환되는 DataFrame 반환.

    N월  시트: 게임매출  — 헤더 행에 "거래처명" 포함
               col D(idx3)=거래처명, col E(idx4)=구분, col F(idx5)=합계
    N월-1 시트: 기타수익 + 용역수익
               col F(idx5)=거래처명, col G(idx6)=구분, col H(idx7)=합계
               구분 예: "기타/로열티수익", "기타/기타", "용역/경영관리수익"
    """
    import openpyxl, re
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception:
        return pd.DataFrame()

    records = []
    seq = 0

    for sheet_name in wb.sheetnames:
        m_game = re.match(r'^(\d+)월$', sheet_name)
        m_etc  = re.match(r'^(\d+)월-1$', sheet_name)
        if not m_game and not m_etc:
            continue

        month_num = int((m_game or m_etc).group(1))
        month_str = f"{month_num}월"
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # 헤더 행 위치 찾기 ("거래처명" 포함 행)
        hdr_row = None
        for i, r in enumerate(rows):
            if any(str(v).strip() == "거래처명" for v in r if v is not None):
                hdr_row = i
                break
        if hdr_row is None:
            continue

        if m_game:
            # ── 게임매출 시트 ───────────────────────────────────────
            for r in rows[hdr_row + 1:]:
                try:
                    vendor = r[3]   # col D
                    sub    = r[4]   # col E
                    amount = r[5]   # col F
                except IndexError:
                    continue
                if not vendor or amount is None:
                    continue
                try:
                    amount = float(amount)
                except (TypeError, ValueError):
                    continue
                if amount == 0:
                    continue

                vendor = str(vendor).strip()
                sub    = str(sub).strip() if sub else "미분류"

                info = VENDOR_MAP.get("4100100", {}).get(vendor)
                if info:
                    display, sub_cat, note = info
                else:
                    display, sub_cat, note = vendor, sub, ""

                records.append(_make_legacy_row(
                    vendor, "4100100", amount,
                    "게임매출", sub_cat, display, month_str, note, seq
                ))
                seq += 1

        else:
            # ── 기타수익 + 용역수익 시트 ───────────────────────────
            for r in rows[hdr_row + 1:]:
                try:
                    vendor  = r[5]   # col F
                    sub_raw = r[6]   # col G
                    amount  = r[7]   # col H
                except IndexError:
                    continue
                if not vendor or amount is None:
                    continue
                try:
                    amount = float(amount)
                except (TypeError, ValueError):
                    continue
                if amount == 0:
                    continue

                vendor  = str(vendor).strip()
                sub_raw = str(sub_raw).strip() if sub_raw else ""

                if sub_raw.startswith("기타/"):
                    cat, acct = "기타수익", "4100500"
                    sub_cat = sub_raw[3:]
                elif sub_raw.startswith("용역/"):
                    cat, acct = "용역수익", "4100300"
                    sub_cat = sub_raw[3:]
                elif sub_raw in ("로열티수익", "기타"):
                    cat, acct = "기타수익", "4100500"
                    sub_cat = sub_raw
                elif sub_raw == "경영관리수익":
                    cat, acct = "용역수익", "4100300"
                    sub_cat = sub_raw
                else:
                    cat, acct = "기타수익", "4100500"
                    sub_cat = sub_raw or "기타"

                info = VENDOR_MAP.get(acct, {}).get(vendor)
                if info:
                    display, _, note = info
                else:
                    display, note = vendor, ""

                records.append(_make_legacy_row(
                    vendor, acct, amount,
                    cat, sub_cat, display, month_str, note, seq
                ))
                seq += 1

    if not records:
        return pd.DataFrame()

    # 컬럼 순서: 원본 30개 + 추가 5개 (grid 파일과 동일한 구조)
    extra_cols = ["매출구분", "세부구분", "표시명", "월", "비고_매핑"]
    df = pd.DataFrame(records, columns=GRID_COLUMNS + extra_cols)
    return df


def append_to_raw_data(existing_df, new_df):
    """
    기존 Raw_Data에 새 데이터를 누적 append.
    동일 (전표번호 + 순번) 중복은 제거.
    """
    if existing_df is None or existing_df.empty:
        return new_df

    combined = pd.concat([existing_df, new_df], ignore_index=True)
    if "전표번호" in combined.columns and "순번" in combined.columns:
        combined = combined.drop_duplicates(subset=["전표번호", "순번"], keep="last")
    return combined.reset_index(drop=True)
